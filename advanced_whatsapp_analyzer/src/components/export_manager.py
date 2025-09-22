"""
Export Manager for Advanced WhatsApp Chat Analyzer
Handles data export in multiple formats (PDF, Excel, CSV, JSON).
"""

import pandas as pd
import json
import io
from datetime import datetime
from typing import Dict, List, Any, Optional, BinaryIO
import logging
from pathlib import Path

# Import optional dependencies
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

logger = logging.getLogger(__name__)

class ExportManager:
    """Manages data export in various formats"""
    
    def __init__(self):
        self.supported_formats = ['csv', 'json']
        
        if EXCEL_AVAILABLE:
            self.supported_formats.append('excel')
        if PDF_AVAILABLE:
            self.supported_formats.append('pdf')
    
    def export_data(self, df: pd.DataFrame, format_type: str, 
                   filename: Optional[str] = None, 
                   include_analysis: bool = True) -> Dict[str, Any]:
        """Export data in specified format"""
        try:
            if format_type not in self.supported_formats:
                return {'error': f'Format {format_type} not supported. Available: {self.supported_formats}'}
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"whatsapp_analysis_{timestamp}"
            
            # Export based on format
            if format_type == 'csv':
                return self._export_csv(df, filename)
            elif format_type == 'excel':
                return self._export_excel(df, filename, include_analysis)
            elif format_type == 'pdf':
                return self._export_pdf(df, filename, include_analysis)
            elif format_type == 'json':
                return self._export_json(df, filename, include_analysis)
            else:
                return {'error': f'Unsupported format: {format_type}'}
                
        except Exception as e:
            logger.error(f"Export failed: {e}")
            return {'error': str(e)}
    
    def _export_csv(self, df: pd.DataFrame, filename: str) -> Dict[str, Any]:
        """Export data as CSV"""
        try:
            # Prepare data for CSV export
            export_df = df.copy()
            
            # Convert datetime columns to strings
            for col in export_df.columns:
                if export_df[col].dtype == 'datetime64[ns]':
                    export_df[col] = export_df[col].dt.strftime('%Y-%m-%d %H:%M:%S')
            
            # Create CSV buffer
            csv_buffer = io.StringIO()
            export_df.to_csv(csv_buffer, index=False, encoding='utf-8')
            
            return {
                'success': True,
                'filename': f"{filename}.csv",
                'data': csv_buffer.getvalue(),
                'size': len(csv_buffer.getvalue()),
                'format': 'csv'
            }
            
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            return {'error': str(e)}
    
    def _export_json(self, df: pd.DataFrame, filename: str, include_analysis: bool = True) -> Dict[str, Any]:
        """Export data as JSON"""
        try:
            # Prepare data
            export_data = {
                'metadata': {
                    'exported_at': datetime.now().isoformat(),
                    'total_messages': len(df),
                    'columns': list(df.columns),
                    'format_version': '2.0'
                },
                'data': []
            }
            
            # Convert DataFrame to records
            df_copy = df.copy()
            
            # Handle datetime columns
            for col in df_copy.columns:
                if df_copy[col].dtype == 'datetime64[ns]':
                    df_copy[col] = df_copy[col].dt.strftime('%Y-%m-%d %H:%M:%S')
            
            export_data['data'] = df_copy.to_dict('records')
            
            # Add analysis if requested
            if include_analysis:
                export_data['analysis'] = self._generate_analysis_summary(df)
            
            # Convert to JSON
            json_data = json.dumps(export_data, indent=2, ensure_ascii=False)
            
            return {
                'success': True,
                'filename': f"{filename}.json",
                'data': json_data,
                'size': len(json_data),
                'format': 'json'
            }
            
        except Exception as e:
            logger.error(f"JSON export failed: {e}")
            return {'error': str(e)}
    
    def _export_excel(self, df: pd.DataFrame, filename: str, include_analysis: bool = True) -> Dict[str, Any]:
        """Export data as Excel workbook with multiple sheets"""
        if not EXCEL_AVAILABLE:
            return {'error': 'Excel export not available - openpyxl not installed'}
        
        try:
            # Create Excel buffer
            excel_buffer = io.BytesIO()
            
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                # Main data sheet
                df_export = df.copy()
                
                # Convert datetime to Excel-friendly format
                for col in df_export.columns:
                    if df_export[col].dtype == 'datetime64[ns]':
                        df_export[col] = pd.to_datetime(df_export[col])
                
                df_export.to_excel(writer, sheet_name='Chat Data', index=False)
                
                # Summary statistics sheet
                if include_analysis:
                    summary_data = self._generate_excel_summary(df)
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # User statistics sheet
                if 'user' in df.columns:
                    user_stats = self._generate_user_statistics(df)
                    user_df = pd.DataFrame(user_stats)
                    user_df.to_excel(writer, sheet_name='User Stats', index=False)
                
                # Daily activity sheet
                daily_activity = df.groupby(df['date'].dt.date).size().reset_index()
                daily_activity.columns = ['Date', 'Message Count']
                daily_activity.to_excel(writer, sheet_name='Daily Activity', index=False)
                
                # Format worksheets
                self._format_excel_sheets(writer)
            
            excel_buffer.seek(0)
            
            return {
                'success': True,
                'filename': f"{filename}.xlsx",
                'data': excel_buffer.getvalue(),
                'size': len(excel_buffer.getvalue()),
                'format': 'excel'
            }
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return {'error': str(e)}
    
    def _export_pdf(self, df: pd.DataFrame, filename: str, include_analysis: bool = True) -> Dict[str, Any]:
        """Export analysis as PDF report"""
        if not PDF_AVAILABLE:
            return {'error': 'PDF export not available - reportlab not installed'}
        
        try:
            # Create PDF buffer
            pdf_buffer = io.BytesIO()
            
            # Create PDF document
            doc = SimpleDocTemplate(
                pdf_buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Get styles
            styles = getSampleStyleSheet()
            title_style = styles['Heading1']
            heading_style = styles['Heading2']
            normal_style = styles['Normal']
            
            # Build story (content)
            story = []
            
            # Title
            story.append(Paragraph("WhatsApp Chat Analysis Report", title_style))
            story.append(Spacer(1, 12))
            
            # Metadata
            story.append(Paragraph("Report Information", heading_style))
            metadata_data = [
                ['Generated Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Total Messages', str(len(df))],
                ['Date Range', f"{df['date'].min().strftime('%Y-%m-%d')} to {df['date'].max().strftime('%Y-%m-%d')}"],
                ['Participants', str(df[df['user'] != 'system']['user'].nunique())],
            ]
            
            metadata_table = Table(metadata_data)
            metadata_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(metadata_table)
            story.append(Spacer(1, 12))
            
            if include_analysis:
                # Analysis summary
                analysis = self._generate_analysis_summary(df)
                
                story.append(Paragraph("Analysis Summary", heading_style))
                story.append(Spacer(1, 6))
                
                # Basic statistics
                if 'basic_stats' in analysis:
                    stats = analysis['basic_stats']
                    stats_data = [
                        ['Metric', 'Value'],
                        ['Average Daily Messages', f"{stats.get('avg_daily_messages', 0):.1f}"],
                        ['Most Active Hour', f"{stats.get('peak_hour', 0)}:00"],
                        ['Weekend Messages %', f"{stats.get('weekend_percentage', 0):.1f}%"],
                    ]
                    
                    if 'avg_message_length' in stats:
                        stats_data.append(['Average Message Length', f"{stats['avg_message_length']:.1f} chars"])
                    
                    stats_table = Table(stats_data)
                    stats_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(stats_table)
                    story.append(Spacer(1, 12))
                
                # User statistics
                if 'user_stats' in analysis:
                    story.append(Paragraph("Top Users", heading_style))
                    story.append(Spacer(1, 6))
                    
                    user_data = [['User', 'Messages', 'Percentage']]
                    for user_info in analysis['user_stats'][:5]:
                        user_data.append([
                            user_info['user'],
                            str(user_info['message_count']),
                            f"{user_info['percentage']:.1f}%"
                        ])
                    
                    user_table = Table(user_data)
                    user_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(user_table)
                    story.append(Spacer(1, 12))
                
                # Sentiment analysis (if available)
                if 'sentiment_stats' in analysis and analysis['sentiment_stats']:
                    story.append(Paragraph("Sentiment Analysis", heading_style))
                    story.append(Spacer(1, 6))
                    
                    sentiment = analysis['sentiment_stats']
                    sentiment_data = [
                        ['Sentiment', 'Percentage'],
                        ['Positive', f"{sentiment.get('positive_pct', 0):.1f}%"],
                        ['Negative', f"{sentiment.get('negative_pct', 0):.1f}%"],
                        ['Neutral', f"{sentiment.get('neutral_pct', 0):.1f}%"],
                    ]
                    
                    sentiment_table = Table(sentiment_data)
                    sentiment_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(sentiment_table)
            
            # Build PDF
            doc.build(story)
            pdf_buffer.seek(0)
            
            return {
                'success': True,
                'filename': f"{filename}.pdf",
                'data': pdf_buffer.getvalue(),
                'size': len(pdf_buffer.getvalue()),
                'format': 'pdf'
            }
            
        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            return {'error': str(e)}
    
    def _generate_analysis_summary(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Generate analysis summary for export"""
        try:
            summary = {}
            
            # Basic statistics
            total_messages = len(df)
            date_range_days = (df['date'].max() - df['date'].min()).days + 1
            
            summary['basic_stats'] = {
                'total_messages': total_messages,
                'date_range_days': date_range_days,
                'avg_daily_messages': total_messages / date_range_days if date_range_days > 0 else 0,
                'unique_users': df[df['user'] != 'system']['user'].nunique(),
                'peak_hour': df.groupby('hour').size().idxmax(),
                'weekend_percentage': len(df[df['is_weekend'] == True]) / len(df) * 100 if 'is_weekend' in df.columns else 0
            }
            
            # Message length stats
            if 'message_length' in df.columns:
                text_df = df[df['message_type'] == 'text']
                summary['basic_stats']['avg_message_length'] = text_df['message_length'].mean()
                summary['basic_stats']['total_words'] = text_df['word_count'].sum() if 'word_count' in text_df.columns else 0
            
            # User statistics
            user_activity = df[df['user'] != 'system']['user'].value_counts()
            summary['user_stats'] = []
            
            for user, count in user_activity.head(10).items():
                summary['user_stats'].append({
                    'user': user,
                    'message_count': count,
                    'percentage': count / total_messages * 100
                })
            
            # Sentiment statistics (if available)
            if 'sentiment_label' in df.columns:
                sentiment_data = df[df['sentiment_label'].notna()]
                if not sentiment_data.empty:
                    sentiment_counts = sentiment_data['sentiment_label'].value_counts()
                    total_sentiment = len(sentiment_data)
                    
                    summary['sentiment_stats'] = {
                        'total_analyzed': total_sentiment,
                        'positive_pct': sentiment_counts.get('positive', 0) / total_sentiment * 100,
                        'negative_pct': sentiment_counts.get('negative', 0) / total_sentiment * 100,
                        'neutral_pct': sentiment_counts.get('neutral', 0) / total_sentiment * 100
                    }
            
            # Activity patterns
            summary['activity_patterns'] = {
                'by_hour': df.groupby('hour').size().to_dict(),
                'by_day': df.groupby('day_name').size().to_dict(),
                'by_time_period': df.groupby('time_period').size().to_dict() if 'time_period' in df.columns else {}
            }
            
            return summary
            
        except Exception as e:
            logger.error(f"Error generating analysis summary: {e}")
            return {}
    
    def _generate_excel_summary(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Generate summary data for Excel export"""
        try:
            summary_data = []
            
            # Basic metrics
            summary_data.extend([
                {'Metric': 'Total Messages', 'Value': len(df)},
                {'Metric': 'Unique Users', 'Value': df[df['user'] != 'system']['user'].nunique()},
                {'Metric': 'Date Range (Days)', 'Value': (df['date'].max() - df['date'].min()).days + 1},
                {'Metric': 'Most Active Hour', 'Value': f"{df.groupby('hour').size().idxmax()}:00"},
            ])
            
            # Content metrics
            if 'message_length' in df.columns:
                text_df = df[df['message_type'] == 'text']
                summary_data.extend([
                    {'Metric': 'Average Message Length', 'Value': f"{text_df['message_length'].mean():.1f} chars"},
                    {'Metric': 'Total Words', 'Value': text_df['word_count'].sum() if 'word_count' in text_df.columns else 'N/A'},
                ])
            
            # Media statistics
            message_types = df['message_type'].value_counts()
            for msg_type, count in message_types.items():
                summary_data.append({
                    'Metric': f'{msg_type.title()} Messages',
                    'Value': count
                })
            
            return summary_data
            
        except Exception as e:
            logger.error(f"Error generating Excel summary: {e}")
            return []
    
    def _generate_user_statistics(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Generate user statistics for Excel export"""
        try:
            user_stats = []
            
            for user in df[df['user'] != 'system']['user'].unique():
                user_df = df[df['user'] == user]
                
                stats = {
                    'User': user,
                    'Total Messages': len(user_df),
                    'Percentage': len(user_df) / len(df) * 100,
                }
                
                if 'message_length' in user_df.columns:
                    text_user_df = user_df[user_df['message_type'] == 'text']
                    stats['Avg Message Length'] = text_user_df['message_length'].mean()
                    stats['Total Words'] = text_user_df['word_count'].sum() if 'word_count' in text_user_df.columns else 0
                
                if 'emoji_count' in user_df.columns:
                    stats['Total Emojis'] = user_df['emoji_count'].sum()
                
                if 'url_count' in user_df.columns:
                    stats['URLs Shared'] = user_df['url_count'].sum()
                
                # Most active time
                stats['Most Active Hour'] = f"{user_df.groupby('hour').size().idxmax()}:00"
                
                user_stats.append(stats)
            
            # Sort by message count
            user_stats = sorted(user_stats, key=lambda x: x['Total Messages'], reverse=True)
            
            return user_stats
            
        except Exception as e:
            logger.error(f"Error generating user statistics: {e}")
            return []
    
    def _format_excel_sheets(self, writer):
        """Format Excel worksheets"""
        try:
            # Get workbook and worksheets
            workbook = writer.book
            
            # Format summary sheet
            if 'Summary' in workbook.sheetnames:
                summary_sheet = workbook['Summary']
                
                # Header formatting
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in summary_sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
            
            # Format user stats sheet
            if 'User Stats' in workbook.sheetnames:
                user_sheet = workbook['User Stats']
                
                for cell in user_sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
            
            # Auto-fit columns
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
                    
        except Exception as e:
            logger.error(f"Error formatting Excel sheets: {e}")
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported export formats"""
        return self.supported_formats.copy()
    
    def validate_export_request(self, df: pd.DataFrame, format_type: str) -> Dict[str, Any]:
        """Validate export request"""
        validation = {'valid': True, 'warnings': [], 'errors': []}
        
        try:
            # Check if format is supported
            if format_type not in self.supported_formats:
                validation['errors'].append(f'Format {format_type} not supported')
                validation['valid'] = False
            
            # Check if DataFrame is empty
            if df.empty:
                validation['errors'].append('No data to export')
                validation['valid'] = False
            
            # Check data size warnings
            if len(df) > 100000:
                validation['warnings'].append('Large dataset - export may take time')
            
            # Format-specific validations
            if format_type == 'excel' and not EXCEL_AVAILABLE:
                validation['errors'].append('Excel export requires openpyxl package')
                validation['valid'] = False
            
            if format_type == 'pdf' and not PDF_AVAILABLE:
                validation['errors'].append('PDF export requires reportlab package')
                validation['valid'] = False
            
        except Exception as e:
            validation['errors'].append(f'Validation error: {str(e)}')
            validation['valid'] = False
        
        return validation